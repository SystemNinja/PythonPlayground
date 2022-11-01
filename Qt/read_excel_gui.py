"""
Ref:
https://doc.qt.io/qtforpython/api.html
https://doc.qt.io/qtforpython/tutorials/basictutorial/uifiles.html#using-ui-files
https://doc.qt.io/qt-6/designer-using-a-ui-file-python.html

Grid Layout:
https://doc.qt.io/qtforpython/PySide6/QtWidgets/QGridLayout.html

File dialog:
https://doc.qt.io/qtforpython/PySide6/QtWidgets/QFileDialog.html

Use 1 function for multiple button signals:
https://stackoverflow.com/questions/15080731/calling-a-function-upon-button-press

Remember last opened folder:
https://www.pythonguis.com/faq/remeber-last-saved-directory-with-qfiledialog/
https://eli.thegreenplace.net/2011/04/25/passing-extra-arguments-to-pyqt-slot/

Update windows cache:
https://realpython.com/python-memcache-efficient-caching/
https://enterprise.arcgis.com/en/server/10.8/publish-services/windows/automating-cache-creation-and-updates-with-geoprocessing.htm
"""

from PySide6 import QtCore, QtWidgets, QtGui
from PySide6.QtWidgets import QDialog,QFileDialog
import sys
import openpyxl

class MainProgram(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.daily_report = ""
        self.weekly_report = ""
        self.root_folder = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports"

        self.label1 = QtWidgets.QLabel("Select daily report file..")
        self.label2 = QtWidgets.QLabel("Select weekly report file..")
        self.browse1 = QtWidgets.QPushButton("Browse")
        self.browse2 = QtWidgets.QPushButton("Browse")
        self.button1 = QtWidgets.QPushButton("Generate Daily Report")
        self.button2 = QtWidgets.QPushButton("Generate Weekly Report")

        layout = QtWidgets.QGridLayout(self)
        layout.addWidget(self.label1, 1, 1)
        layout.addWidget(self.label2, 2, 1)
        layout.addWidget(self.browse1, 1, 2)
        layout.addWidget(self.browse2, 2, 2)
        layout.addWidget(self.button1, 4, 1)
        layout.addWidget(self.button2, 4, 2)

        self.browse1.clicked.connect(self.browse_file)
        self.button1.clicked.connect(self.process_daily_report)

    def browse_file(self):
        # fname = QFileDialog.getOpenFileName(self,"Select file",self.root_folder,"Excel Files (*.xlsx *.csv)")
        fname = ""
        dialog = QFileDialog(self)
        dialog.setDirectory(self.root_folder)
        dialog.setFileMode(QFileDialog.ExistingFile)
        dialog.setNameFilter("Excel file (*.xlsx *.csv)")
        dialog.setViewMode(QFileDialog.Detail)
        if dialog.exec():
            fname = dialog.selectedFiles()

        if fname == "":
            self.label1.setText("Select file..")
        else:
            self.label1.setText(fname[0])
            self.daily_report = fname[0]


    def process_daily_report(self):
        backup_report = self.daily_report
        weekly_report = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports\Weekly\Backup Detail Report_Setton_Thursday, August 25, 2022.xlsx"
        backup_template = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports\Daily_Report_Template.xlsx"
        report_workbook = openpyxl.load_workbook(filename=backup_report)
        template_workbook = openpyxl.load_workbook(filename=backup_template)
        report_sheet = report_workbook.active
        template_sheet = template_workbook.active
        row_count = report_sheet.max_row
        success = 0
        warning = 0
        failure = 0
        cell_success = template_sheet["B2"]
        cell_warning = template_sheet["B3"]
        cell_failure = template_sheet["B4"]

        for value in report_sheet.iter_rows(min_row=2, max_row=row_count, min_col=8, max_col=8, values_only=True):
            value = str(value).replace("'", "").replace("(", "").replace(")", "").replace(",", "")
            if (value == "Success"):
                success = success + 1
            elif (value == "Warning"):
                warning += 1
            elif (value == "Failure"):
                failure += 1
            else:
                continue

        # print(f"Success: {success}\n")
        # print(f"Warning: {warning}\n")
        # print(f"Failure: {failure}\n")

        cell_success.value = success
        cell_warning.value = warning
        cell_failure.value = failure
        template_workbook.save(backup_template)
        template_workbook.close()
        report_workbook.close()

if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    main_window = MainProgram()
    main_window.setWindowTitle("Process Report")
    #main_window.setFixedSize(300, 200)
    main_window.show()

    sys.exit(app.exec())

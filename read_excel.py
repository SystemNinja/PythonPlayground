import openpyxl

backup_report = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports\Daily\Backup Detail Report_Setton_Thursday, September 1, 2022.xlsx"
weekly_report = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports\Weekly\Backup Detail Report_Setton_Thursday, August 25, 2022.xlsx"
backup_template = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports\Daily_Report_Template.xlsx"
report_workbook = openpyxl.load_workbook(filename=backup_report)
template_workbook = openpyxl.load_workbook(filename=backup_template)
report_sheet = report_workbook.active
template_sheet = template_workbook.active
row_count = report_sheet.max_row
success = 0
warning = 0
failure = 0
cell_success = template_sheet["B2"]
cell_warning = template_sheet["B3"]
cell_failure = template_sheet["B4"]

for value in report_sheet.iter_rows(min_row=2,max_row=row_count,min_col=8,max_col=8,values_only=True):
    value=str(value).replace("'","").replace("(","").replace(")","").replace(",","")
    if (value == "Success"):
        success = success + 1
    elif(value == "Warning"):
        warning += 1
    elif(value == "Failure"):
        failure += 1
    else:
        continue

print(f"Success: {success}\n")
print(f"Warning: {warning}\n")
print(f"Failure: {failure}\n")

cell_success.value = success
cell_warning.value = warning
cell_failure.value = failure
template_workbook.save(backup_template)
template_workbook.close()
report_workbook.close()
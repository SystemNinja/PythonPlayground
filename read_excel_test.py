"""
https://stackoverflow.com/questions/13377793/is-it-possible-to-get-an-excel-documents-row-count-without-loading-the-entire-d
https://realpython.com/openpyxl-excel-spreadsheets-python/
https://linuxconfig.org/how-to-manipulate-excel-spreadsheets-with-python-and-openpyxl
"""

import openpyxl
import json

file_name = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports\Daily\Backup Detail Report_Setton_Thursday, July 14, 2022.xlsx"
file_name2 = r"C:\Users\white\OneDrive - Setton Consulting Inc\Backupradar\Reports\Daily\Backup Detail Report_Setton_Monday, July 18, 2022.xlsx"
workbook = openpyxl.load_workbook(filename=file_name2)
# sheet = workbook.worksheets[0]
sheet = workbook.active
row_count = sheet.max_row

print(row_count)
print(sheet.cell(row=row_count, column=8).value)

# List values of all columns
for value in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
    print(value)

items = { }
for value in sheet.iter_rows(min_col=7,max_col=10, values_only=True):
    item_id = value[0]
    items = {
        "test1": value[1],
        "test2": value[2],
        "test3": value[3]
    }
    items[item_id] = items

print(items.get("test1"))
# print("Get all cells from column.")
# print(sheet["J"])
#
# print("Iterate through entire dataset")
# for row in sheet.rows:
#     print(row)
